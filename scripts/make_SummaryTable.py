#! /usr/bin/env python

# This script loops through all the samples,
# creates a summary stats table
import csv
import os
from argparse import ArgumentParser
from pypiper import AttributeDict
import yaml


# Argument Parsing
# #######################################################################################
parser = ArgumentParser(description='make_SummaryTable')
parser.add_argument('-c', '--config-file', dest='config_file', help="path to YAML config file", required=True, type=str)
parser.add_argument('--excel', dest='excel', action='store_true', help="generate extra XLS and XLSX sheet", default=False, required=False)
# Charles : On time the legacy/rigid mode will be removed
parser.add_argument('--rigid', dest='rigid', action='store_true', help="the legacy rigid mode that only takes in the hard-coded values", default=False, required=False)
args = parser.parse_args()

with open(args.config_file, 'r') as config_file:
	config_yaml = yaml.load(config_file)
	config = AttributeDict(config_yaml, default=True)
paths = config.paths



if not os.path.exists(paths.output_dir):
	raise Exception(paths.output_dir + " : project directory does not exist!")


# FOR RIGID
# #######################################################################################
fields_in = []
fields_out = []
if args.rigid:
	# the hard-coded fields for the legacy/rigid mode
	fields_in = ['sample_name','instrument_model','flowcell','lane','read_length','Read_type','organism','Genome'\
	,'cell_type','Raw_reads','Trimmed_reads','Trimmed_rate','Aligned_reads','Aligned_rate'\
	,'Multimap_reads','Multimap_rate','Unique_CpGs','Total_CpGs','meanCoverage',\
	'bisulfiteConversionRate','globalMethylationMean',\
	'K1_unmethylated_count','K1_unmethylated_meth','K3_methylated_count','K3_methylated_meth']
	fields_out = ['Sample','Instrument','Flowcell','Lane','Read Length','Read Type','Organism','Genome'\
	,'Cell Type','Raw Reads','Trimmed Reads','Trimmed Rate','Aligned Reads','Aligned Rate'\
	,'Multimap Reads','Multimap Rate','Unique CpGs','Total CpGs','Mean Coverage',\
	'Bisulfite Conversion Rate',' Global Methylation Mean',\
	'K1 Unmethylated Count','K1 Unmethylated Meth','K3 Methylated Count','K3 Methylated Meth']


# Open samples CSV file
# #######################################################################################
csv_file_path = os.path.join(os.path.dirname(args.config_file),config.metadata.sample_annotation)
print("\nOpening CSV file: " + csv_file_path)
if os.path.isfile(csv_file_path):
	csv_file = open(os.path.join(os.path.dirname(args.config_file),config.metadata.sample_annotation), 'rb')
	print("Found " + csv_file_path)
else:
	raise Exception(csv_file_path + " : that file does not exist!")
csv_reader = csv.DictReader(csv_file)


# Looping over all samples
# #######################################################################################
global_list = dict()
global_keys = dict()

pipelines = []
sample_count = 0
column_count = 0
print("\nStart iterating over samples")

for row in csv_reader:

	sample_count += 1
	sample_name = row['sample_name']
	print("\n##### Processing sample #"+ str(sample_count) + " : " + sample_name + " #####")

	# wrap this all in a try block, so it can skip a few bad samples
	# without breaking the whole thing 
	try:

		# Open sample TSV stat file
		stat_file_dir = os.path.join(paths.output_dir,paths.results_subdir,sample_name)
		stat_file_path = os.path.join(paths.output_dir,paths.results_subdir,sample_name,row['library']+'_stats.tsv')
		if not os.path.isfile(stat_file_path):
			for thefile in os.listdir(stat_file_dir):
				if thefile.endswith("stats.tsv"): stat_file_path = os.path.join(stat_file_dir,thefile)
		if os.path.isfile(stat_file_path):
			stat_file = open(stat_file_path, 'rb')
			print("Found: " + stat_file_path)
		else:
			raise Exception(stat_file_path + " : file does not exist!")


		stats_dict = dict()
		stats_dict_keys = dict()

		# Check if file has third column -> define pipelines based on that
		# plus read info from file

		for line in stat_file:
			line_content = line.split('\t')
			key = line_content[0]
			value = line_content[1]
			pip = "x"
			if len(line_content) == 3: 
				pip = line_content[2].strip()
			pipelines.append(pip)
			if not pip in stats_dict: stats_dict[pip] = dict()
			if not pip in stats_dict_keys: stats_dict_keys[pip] = []
			stats_dict[pip][key] = value.strip()
			stats_dict_keys[pip].append(key)

		pipelines = list(set(pipelines))
		print "Pipelines: " + str(pipelines)


		# stats_dict and stats_dict_keys are pipeline specific
		for pip in pipelines:
			if not pip in global_list: global_list[pip] = []
			if not pip in global_keys: global_keys[pip] = []



		# if there are two pipelines make sure that certain values are present in both
		missing_cols = ["Raw_reads", "Fastq_reads", "Trimmed_reads", "Trim_loss_rate"]
		if len(pipelines) == 2:
			for col in missing_cols:
				if not col in stats_dict[pipelines[1]] and col in stats_dict[pipelines[0]]: stats_dict[pipelines[1]][col] = stats_dict[pipelines[0]][col]
				if not col in stats_dict[pipelines[0]] and col in stats_dict[pipelines[1]]: stats_dict[pipelines[0]][col] = stats_dict[pipelines[1]][col]
				for pip in pipelines:
					stats_dict_keys[pip] = list(set(stats_dict_keys[pip] + missing_cols))

		# Write to global list and keys
		new_row = dict()
		column_count = 0
		for pip in pipelines:
			new_row = row.copy()
			new_row.update(stats_dict[pip])
			global_list[pip].append(new_row)
			global_keys[pip] = csv_reader.fieldnames + stats_dict_keys[pip]


	except Exception as e:

		print("Sample " + sample_name + " failed. Error: " + str(e))

csv_file.close()
# print global_keys
# print global_list

# Writing to Output Files
# #######################################################################################
if not args.rigid:
	
	# Writing TSV file
	# #######################################################################################

	for pip in pipelines:
		
		pip_nam = "_" + pip
		if pip_nam == "_x": pip_nam = ""
		tsv_outfile_path = os.path.join(paths.output_dir,os.path.basename(paths.output_dir)+ pip_nam + '_stats_summary.tsv')
		tsv_outfile = open(tsv_outfile_path, 'w')
	
		if global_list[pip] and global_keys[pip]:

			tsv_writer = csv.DictWriter(tsv_outfile, fieldnames=global_keys[pip], delimiter='\t')
			tsv_writer.writeheader()

			for i,sample in enumerate(global_list[pip]):
				tsv_writer.writerow(sample)
				if args.excel:
					for j,field in enumerate(global_keys[pip]):
						if i == 0: xls_sheet.write(0, j, field)
						xls_sheet.write(i+1, j, sample[field])

		tsv_outfile.close()

	print("\nInput used: " + csv_file_path)
	print("Results TSV file: " + tsv_outfile_path)




	# Output XLS file
	# #######################################################################################
	if args.excel:
		
		raise Exception("--excel not implemented")
			
		import xlwt
		
		for pip in pipelines:
			
			pip_nam = "_" + pip
			if pip_nam == "_x": pip_nam = ""

			xls_book = xlwt.Workbook(encoding="utf-8")
			xls_sheet_name = "Stats" + pip_nam
			xls_sheet = xls_book.add_sheet(xls_sheet_name)

			# Where should this be written? Here or below?
			# if args.rigid:
			# 	for i,field in enumerate(fields_out):
			# 		xls_sheet.write(0, i, field)

			import xlrd
			import openpyxl

			# saving the XLS sheet
			xls_filename = os.path.join(paths.output_dir,os.path.basename(paths.output_dir)+'_stats_summary.xls')
			xls_book.save(xls_filename)
			print("Results XLS file: " + xls_filename)

			# convert XLS sheet to XLSX format
			xlsx_book_in = xlrd.open_workbook(xls_filename)
			index = 0
			nrows = sample_count + 2
			ncols = 0
			if global_keys[pip]: ncols = len(global_keys[pip])
			else: ncols = column_count
			ncols += 1
			xlsx_sheet_in = xlsx_book_in.sheet_by_index(0)
			xlsx_book_out = openpyxl.Workbook()
			xlsx_sheet = xlsx_book_out.active
			xlsx_sheet.title = xls_sheet_name
			for row in range(1, nrows):
				for col in range(1, ncols):
					xlsx_sheet.cell(row=row, column=col).value = xlsx_sheet_in.cell_value(row-1, col-1)
			xlsx_filename = os.path.join(paths.output_dir,os.path.basename(paths.output_dir)+'_stats_summary.xlsx')
			xlsx_book_out.save(xlsx_filename)
			print("Results XLSX file: " + xlsx_filename)

	print("\n")



# RIGID		
else:
	if args.excel:
		raise Exception("--excel not implemented for option --rigid")
		
	for pip in pipelines:

		pip_nam = "_" + pip
		if pip_nam == "_x": pip_nam = ""
		# Open file to write to
		tsv_outfile_path = os.path.join(paths.output_dir,os.path.basename(paths.output_dir)+ pip_nam + '_stats_summary.tsv')
		tsv_outfile = open(tsv_outfile_path, 'w')
		tsv_writer = csv.DictWriter(tsv_outfile, fieldnames=fields_out, delimiter='\t')
		tsv_writer.writeheader()
		
		

		
		# for each sample data (one element of the global list)
		for sample_dict in global_list[pip]:
			
			new_row = dict()
			# Write each field
			for i in range(0,len(fields_in)):

				field = fields_in[i]
				field_out = fields_out[i]
				content = str('')
				content_float = float(-1e10)
				content_int = int(-1)

				# extract all the listed fields
				# some data types might not have all the fields in stats_dict, then catch the KeyError
				try:
					if field == 'Trimmed_rate':
						content_float = 100.0*float(sample_dict['Trimmed_reads'])/float(sample_dict['Raw_reads'])
					elif field == 'Aligned_rate':
						content_float = 100.0*float(sample_dict['Aligned_reads'])/float(sample_dict['Trimmed_reads'])
					elif field == 'Multimap_rate':
						content_float = 100.0*float(sample_dict['Multimap_reads'])/float(sample_dict['Trimmed_reads'])
					elif sample_dict.has_key(field):
						content = str(sample_dict[field].strip())
					else:
						content = 'NA'
						print("No field called: " + field)
				except KeyError:
					content = 'NA'
					print("Data missing to calculate: " + field)

				# convert str to float or int if needed
				got_comma = content.find('.')
				try:
					content_float = float(content)
				except ValueError:
					pass
				if not got_comma:
					content_int = int(content_float)

				# write the field for each row
				if content_int > -1:
					column_count += 1
					new_row[field_out] = content_int
					if args.excel: xls_sheet.write(sample_count, i, content_int)
				elif content_float > -1e10:
					column_count += 1
					new_row[field_out] = content_float
					if args.excel: xls_sheet.write(sample_count, i, content_float)
				else:
					column_count += 1
					new_row[field_out] = content
					if args.excel: xls_sheet.write(sample_count, i, content)
			
			tsv_writer.writerow(new_row)
		
	tsv_outfile.close()





